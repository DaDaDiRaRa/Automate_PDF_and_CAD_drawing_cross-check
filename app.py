"""
app.py  —  PDF <-> CAD Drawing Cross-Check  (V8)
================================================
Compares a master "PDF Drawing List" against the actual CAD drawings
(.dwg / .dxf) in a directory, without moving any files so Xref relative
paths remain intact.

Pipeline
--------
1. Interactive prompts  : TARGET_DIR, PDF_PATH, BLOCK_NAME
2. PDF extraction       : pdfplumber with 3-tier fallback
                            Tier 1  – default table extraction
                            Tier 2  – tuned table settings (text strategy)
                            Tier 3  – raw text + per-line regex parsing
3. DWG extraction       : ezdxf, Model Space only, os.chdir() preserves Xref
4. Compare & Report     : outer-merge on Drawing Number -> report.xlsx
                          red cells on mismatch / missing rows

---- V8 changes ----
* ODA_PATH now defaults to the standard Windows install of OdaFileConverter
  27.1.0.  _configure_odafc() is called once before the first DWG open.
* _load_doc() wraps odafc.readfile in try/except and emits a clean
  "ODA Converter not found at [...]" message on failure.
* Drawing-number parsing now handles the spaced format "AA - 401" that
  appears in real-world Korean drawing lists, and every extracted number
  is normalised through _normalize_drawing_number() so the PDF and DWG
  sides always merge on the canonical "AA-401" form.
"""

from __future__ import annotations

import glob
import os
import re
import sys
import traceback
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
import pdfplumber
import ezdxf
from ezdxf import bbox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ============================================================================
# GLOBAL TWEAKABLE CONSTANTS
# ============================================================================

# --- ODA File Converter ------------------------------------------------------
# Full path to the OdaFileConverter executable.
#
# Leave empty ("") to enable automatic detection. On startup the script will:
#   1. search common install roots (C:\Program Files\ODA\*  and
#      C:\Program Files (x86)\ODA\*) for OdaFileConverter.exe, picking the
#      newest version folder it finds, then
#   2. if nothing is found, prompt you interactively for the full path.
#
# You can still hardcode a path here if you prefer to skip detection, e.g.:
#   ODA_PATH = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"
ODA_PATH: str = ""

# --- Title-block search ratios ----------------------------------------------
X_RATIO: float = 0.101    # 10.10 %  search width  = Title-Block Width  * X_RATIO
Y_RATIO: float = 0.2138   # 21.38 %  search height = Title-Block Height * Y_RATIO

# --- Misc --------------------------------------------------------------------
REPORT_NAME: str = "report.xlsx"

# --- Regexes -----------------------------------------------------------------
# Short alphanumeric token that looks like a drawing number (fallback parser).
# Examples: A-101  S_002  MEP-12B  DWG01
_DRAWING_NUMBER_RE = re.compile(
    r"^[A-Za-z]{0,5}[-_]?\d{1,5}[A-Za-z0-9\-_.]*$"
)

# Scale expressions at the tail of a text line.
_SCALE_RE = re.compile(
    r"\b(1\s*[:/]\s*\d+|NTS|N\.T\.S\.|AS\s+SHOWN)\b",
    re.IGNORECASE,
)

# Primary line-level regex for real-world Korean drawing lists of the form:
#     "AA - 401 101동 영구저류조 (SL+53.355) 전체 평면도 1/100 1/200"
# Captures:
#     1) letter prefix      : "AA"
#     2) numeric core       : "401"      (optionally followed by A-Z0-9)
#     3) drawing name       : "101동 ... 전체 평면도"       (non-greedy)
#     4) trailing scale(s)  : "1/100 1/200"  (optional - may be absent)
_PDF_LINE_SPACED_RE = re.compile(
    r"^\s*([A-Za-z]{1,5})\s*-\s*(\d{1,5}[A-Za-z0-9]*)\s+"
    r"(.+?)"
    r"(?:\s+("
    r"\d+\s*/\s*\d+(?:\s+\d+\s*/\s*\d+)*"        # 1/100  or  1/100 1/200
    r"|\d+\s*:\s*\d+"                              # 1:100
    r"|NTS|N\.T\.S\.|AS\s+SHOWN"                   # textual scales
    r"))?\s*$",
    re.IGNORECASE,
)


def _normalize_drawing_number(s: str) -> str:
    """
    Strip all internal whitespace from a drawing number so that
    "AA - 401" and "AA-401" compare equal when merging PDF vs DWG rows.
    """
    if not s:
        return ""
    return re.sub(r"\s+", "", s)


# ============================================================================
# 1.  INTERACTIVE CLI PROMPTS
# ============================================================================
def _prompt_path(label: str, *, must_be_dir: bool = False,
                 must_be_file: bool = False) -> str:
    """Prompt the user for a filesystem path, expand it, and validate."""
    while True:
        raw = input(label).strip().strip('"').strip("'")
        if not raw:
            print("    ! Empty input. Please try again.")
            continue
        path = os.path.expanduser(os.path.expandvars(raw))
        if must_be_dir and not os.path.isdir(path):
            print(f"    ! Not a valid directory: {path}")
            continue
        if must_be_file and not os.path.isfile(path):
            print(f"    ! Not a valid file: {path}")
            continue
        return os.path.abspath(path)


def prompt_inputs() -> Tuple[str, str, str]:
    """Collect TARGET_DIR, PDF_PATH and BLOCK_NAME from the user."""
    print("=" * 72)
    print(" PDF <-> CAD Drawing Cross-Check")
    print("=" * 72)

    target_dir = _prompt_path(
        "Enter the full directory path that contains the DWG files: ",
        must_be_dir=True,
    )
    pdf_path = _prompt_path(
        "Enter the full path of the master PDF drawing list: ",
        must_be_file=True,
    )
    block_name = input("Enter the Name of the Title Block to search for: ").strip()
    if not block_name:
        print("[ERROR] Title block name cannot be empty.")
        sys.exit(1)

    return target_dir, pdf_path, block_name


# ============================================================================
# 2.  PDF EXTRACTION  (3-tier fallback)
# ============================================================================

# ---- helpers ----------------------------------------------------------------

def _find_col(header: List[str], keys: List[str]) -> Optional[int]:
    """Return the column index of the first header cell containing any key."""
    for i, cell in enumerate(header):
        for k in keys:
            if k in cell:
                return i
    return None


def _debug_pdf_page(page) -> None:
    """Print a raw-text snippet from a page so you can diagnose parsing."""
    raw = page.extract_text() or ""
    snippet = raw[:800]
    print("[PDF ] -------- raw text snippet (page 1) --------")
    for line in snippet.splitlines():
        print(f"[PDF ]   {line}")
    print("[PDF ] -------------------------------------------------")


def _rows_from_table(table: list) -> List[dict]:
    """Convert a pdfplumber table (list-of-lists) into drawing-list dicts."""
    if not table or len(table) < 2:
        return []
    header = [(c or "").strip().lower() for c in table[0]]
    idx_no    = _find_col(header, ["drawing number", "dwg no", "dwg. no",
                                   "doc no", "doc. no", "no."])
    idx_name  = _find_col(header, ["drawing name", "drawing title", "title",
                                   "description", "name"])
    idx_scale = _find_col(header, ["scale"])
    if idx_no is None or idx_name is None:
        return []
    rows: List[dict] = []
    for raw in table[1:]:
        if not raw or all((c or "").strip() == "" for c in raw):
            continue
        number = (raw[idx_no]   or "").strip() if idx_no   < len(raw) else ""
        name   = (raw[idx_name] or "").strip() if idx_name < len(raw) else ""
        scale  = ""
        if idx_scale is not None and idx_scale < len(raw):
            scale = (raw[idx_scale] or "").strip()
        if not number:
            continue
        rows.append({
            "Drawing Number": _normalize_drawing_number(number),
            "Drawing Name":   name,
            "Scale":          scale,
        })
    return rows


# ---- extraction tiers -------------------------------------------------------

def _try_table_default(page) -> List[dict]:
    """Tier 1: pdfplumber default table detection."""
    rows: List[dict] = []
    for table in (page.extract_tables() or []):
        rows.extend(_rows_from_table(table))
    return rows


def _try_table_tuned(page) -> List[dict]:
    """Tier 2: text-strategy table detection — works on borderless tables."""
    settings = {
        "vertical_strategy":    "text",
        "horizontal_strategy":  "text",
        "snap_tolerance":       5,
        "join_tolerance":       3,
        "edge_min_length":      10,
        "min_words_vertical":   1,
        "min_words_horizontal": 1,
    }
    rows: List[dict] = []
    try:
        for table in (page.extract_tables(table_settings=settings) or []):
            rows.extend(_rows_from_table(table))
    except Exception:
        pass
    return rows


def _parse_text_line(line: str) -> Optional[dict]:
    """
    Parse one raw text line into {Drawing Number, Drawing Name, Scale}.

    Two strategies:

    A) Spaced-number regex - handles "AA - 401 Name ... 1/100":
           AA - 401 101동 영구저류조 (SL+53.355) 전체 평면도 1/100 1/200

    B) Fixed-width fallback - splits on 2+ spaces and uses the short
       alphanumeric heuristic for the first column.
    """
    line = line.strip()
    if not line:
        return None

    # ---- Strategy A: spaced drawing number regex ---------------------------
    m = _PDF_LINE_SPACED_RE.match(line)
    if m:
        letters, digits, name, scale = m.groups()
        return {
            "Drawing Number": _normalize_drawing_number(f"{letters}-{digits}"),
            "Drawing Name":   (name or "").strip(),
            "Scale":          (scale or "").strip(),
        }

    # ---- Strategy B: fixed-width fallback ----------------------------------
    scale = ""
    m_sc = _SCALE_RE.search(line)
    if m_sc:
        scale = m_sc.group(1).strip()
        line = (line[:m_sc.start()] + line[m_sc.end():]).strip()

    parts = re.split(r"\s{2,}", line)
    if len(parts) < 2:
        return None

    candidate_no   = parts[0].strip()
    candidate_name = " ".join(p.strip() for p in parts[1:] if p.strip())

    squashed = candidate_no.replace(" ", "")
    if _DRAWING_NUMBER_RE.match(squashed) and any(ch.isdigit() for ch in squashed):
        return {
            "Drawing Number": _normalize_drawing_number(candidate_no),
            "Drawing Name":   candidate_name,
            "Scale":          scale,
        }
    return None


def _try_text_regex(page) -> List[dict]:
    """Tier 3: raw text extraction followed by per-line regex parsing."""
    raw = page.extract_text() or ""
    rows: List[dict] = []
    for line in raw.splitlines():
        row = _parse_text_line(line)
        if row:
            rows.append(row)
    return rows


# ---- main PDF function ------------------------------------------------------

def extract_pdf_table(pdf_path: str) -> pd.DataFrame:
    """
    Parse every page of *pdf_path* and return a DataFrame with columns
    ``Drawing Number``, ``Drawing Name``, ``Scale``.
    """
    print(f"[PDF ] Opening: {pdf_path}")
    all_rows: List[dict] = []
    debug_done = False

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_no, page in enumerate(pdf.pages, 1):

                if not debug_done:
                    _debug_pdf_page(page)
                    debug_done = True

                before = len(all_rows)

                page_rows = _try_table_default(page)
                if page_rows:
                    strategy = "table/default"
                else:
                    page_rows = _try_table_tuned(page)
                    if page_rows:
                        strategy = "table/tuned"
                    else:
                        page_rows = _try_text_regex(page)
                        strategy = "text+regex"

                all_rows.extend(page_rows)
                added = len(all_rows) - before
                print(f"[PDF ] Page {page_no}: +{added} rows  "
                      f"[strategy: {strategy}]  (total {len(all_rows)})")

    except FileNotFoundError:
        print(f"[ERROR] PDF not found: {pdf_path}")
        raise
    except Exception as exc:
        print(f"[ERROR] Failed to read PDF: {exc}")
        raise

    df = pd.DataFrame(all_rows, columns=["Drawing Number", "Drawing Name", "Scale"])
    if not df.empty:
        df = df.drop_duplicates(subset=["Drawing Number"]).reset_index(drop=True)
    print(f"[PDF ] Extracted {len(df)} unique drawings")
    return df


# ============================================================================
# 3.  DWG / DXF EXTRACTION  (Model Space only)
# ============================================================================

# --- ODA File Converter configuration ---------------------------------------
_odafc_configured: bool = False
_oda_resolved_path: str = ""   # populated by _resolve_oda_path()

# Standard install roots searched during auto-detection.
_ODA_SEARCH_ROOTS: Tuple[str, ...] = (
    r"C:\Program Files\ODA",
    r"C:\Program Files (x86)\ODA",
)
# Executable filenames (case variants — NTFS is case-insensitive but the
# glob itself on other platforms is not).
_ODA_EXE_NAMES: Tuple[str, ...] = (
    "ODAFileConverter.exe",
    "OdaFileConverter.exe",
)


def _autodetect_oda_path() -> str:
    """
    Walk the common ODA install roots and return the full path to
    OdaFileConverter.exe (preferring the newest version folder). Returns
    an empty string if nothing is found.
    """
    candidates: List[str] = []
    for root in _ODA_SEARCH_ROOTS:
        if not os.path.isdir(root):
            continue
        for exe in _ODA_EXE_NAMES:
            # Typical layout: <root>\ODAFileConverter <version>\<exe>
            candidates.extend(glob.glob(os.path.join(root, "*", exe)))
            # Also accept the executable placed directly in <root>\<exe>
            candidates.extend(glob.glob(os.path.join(root, exe)))

    # De-duplicate and pick the lexicographically-highest match — for a
    # folder like "ODAFileConverter 27.1.0" this naturally prefers the
    # most recent version.
    candidates = sorted({os.path.normpath(c) for c in candidates}, reverse=True)

    for c in candidates:
        if os.path.isfile(c):
            return c
    return ""


def _resolve_oda_path() -> str:
    """
    Decide which OdaFileConverter.exe to use, in order of preference:

      1. Honour ODA_PATH if the user set it manually at the top of this file.
      2. Auto-detect inside the standard Windows install roots.
      3. Fall back to interactive input().

    The final value is cached in ``_oda_resolved_path`` and returned.
    """
    global _oda_resolved_path

    # Hardcoded override wins.
    if ODA_PATH and os.path.isfile(ODA_PATH):
        _oda_resolved_path = ODA_PATH
        return _oda_resolved_path
    if ODA_PATH and not os.path.isfile(ODA_PATH):
        print(f"[WARN] ODA_PATH is set but not a valid file: {ODA_PATH}")

    # Auto-detect.
    found = _autodetect_oda_path()
    if found:
        print(f"[INFO] Auto-detected ODA Converter at: {found}")
        _oda_resolved_path = found
        return _oda_resolved_path

    # Fallback: prompt the user.
    print("[WARN] Could not auto-detect OdaFileConverter.exe in the standard")
    print("       install roots:")
    for r in _ODA_SEARCH_ROOTS:
        print(f"           {r}")
    while True:
        raw = input("Enter the exact file path for OdaFileConverter.exe: ").strip().strip('"').strip("'")
        if not raw:
            print("    ! Empty input. Please try again.")
            continue
        path = os.path.expanduser(os.path.expandvars(raw))
        if not os.path.isfile(path):
            print(f"    ! Not a valid file: {path}")
            continue
        _oda_resolved_path = os.path.abspath(path)
        return _oda_resolved_path


def _configure_odafc() -> None:
    """
    Resolve the OdaFileConverter executable path and push it into ezdxf's
    ``odafc`` addon so subsequent DWG opens succeed. Called lazily (from
    ``_load_doc``) on the first DWG file encountered.

    Three things are done in order so the setting sticks across every
    ezdxf variant we have seen in the wild:

      1. ``ezdxf.addons.odafc.configs.odafc_exec_path``  (configs submodule
         — present in some forks, silently ignored if missing)
      2. ``ezdxf.addons.odafc.win_exec_path`` /
         ``ezdxf.addons.odafc.unix_exec_path``            (documented in 1.x)
      3. Inject the binary's directory into ``os.environ["PATH"]`` so that
         the ``shutil.which()`` lookup inside ezdxf finds the executable
         even when the above two attributes are ignored. This is the
         actual workaround — several ezdxf releases hard-check PATH via
         shutil.which() before honouring the configured exec_path.
    """
    global _odafc_configured
    if _odafc_configured:
        return

    path = _resolve_oda_path()
    if not path:
        # Nothing resolved - let ezdxf try its own PATH search as last resort.
        _odafc_configured = True
        return

    # --- (1) configs submodule (try first, gracefully skip if missing) -----
    try:
        import ezdxf.addons.odafc.configs as odafc_configs  # type: ignore
        odafc_configs.odafc_exec_path = path
    except Exception:
        pass

    # --- (2) documented module-level attribute API --------------------------
    try:
        from ezdxf.addons import odafc
        if sys.platform == "win32":
            odafc.win_exec_path = path
        else:
            odafc.unix_exec_path = path
    except Exception as exc:
        print(f"[WARN] Could not set ezdxf odafc exec path: {exc}")

    # --- (3) inject the binary's directory into os.environ["PATH"] ---------
    # This is the real workaround: ezdxf's odafc helper calls shutil.which()
    # under the hood, which only consults the PATH environment variable. By
    # prepending the folder that contains OdaFileConverter.exe we guarantee
    # shutil.which() returns a hit regardless of which ezdxf release is used.
    exe_dir = os.path.dirname(path)
    if exe_dir:
        current_path = os.environ.get("PATH", "")
        path_parts = current_path.split(os.pathsep) if current_path else []
        if exe_dir not in path_parts:
            os.environ["PATH"] = exe_dir + os.pathsep + current_path
            print(f"[INFO] Injected ODA directory into PATH: {exe_dir}")

    _odafc_configured = True


def _load_doc(path: Path):
    """Load a DXF natively or a DWG via the ODA File Converter addon."""
    suffix = path.suffix.lower()

    if suffix == ".dxf":
        try:
            return ezdxf.readfile(str(path))
        except Exception as exc:
            raise RuntimeError(f"ezdxf could not read {path.name}: {exc}") from exc

    if suffix == ".dwg":
        _configure_odafc()
        shown_path = _oda_resolved_path or ODA_PATH or "<system PATH>"
        try:
            from ezdxf.addons import odafc
            return odafc.readfile(str(path))
        except FileNotFoundError as exc:
            raise RuntimeError(
                f"ODA Converter not found at [{shown_path}] "
                f"(underlying error: {exc})"
            ) from exc
        except Exception as exc:
            msg = str(exc).lower()
            if "odafileconverter" in msg or "oda_file_converter" in msg \
               or "no such file" in msg or "cannot find" in msg:
                raise RuntimeError(
                    f"ODA Converter not found at [{shown_path}] "
                    f"(underlying error: {exc})"
                ) from exc
            raise RuntimeError(
                f"odafc failed to read {path.name}: {exc}"
            ) from exc

    raise ValueError(f"Unsupported CAD file extension: {suffix}")


def _entity_point(ent) -> Optional[Tuple[float, float]]:
    """Return a representative (x, y) insertion point for TEXT / MTEXT."""
    t = ent.dxftype()
    try:
        if t == "TEXT":
            if getattr(ent.dxf, "halign", 0) or getattr(ent.dxf, "valign", 0):
                p = ent.dxf.align_point
            else:
                p = ent.dxf.insert
            return (float(p[0]), float(p[1]))
        if t == "MTEXT":
            p = ent.dxf.insert
            return (float(p[0]), float(p[1]))
    except Exception:
        return None
    return None


def _entity_text(ent) -> str:
    """Return the plain text content of a TEXT / MTEXT entity."""
    t = ent.dxftype()
    try:
        if t == "TEXT":
            return (ent.dxf.text or "").strip()
        if t == "MTEXT":
            return ent.plain_text().strip()
    except Exception:
        return ""
    return ""


def _compute_insert_bbox(insert):
    """Return the bounding box of an INSERT or None if unavailable."""
    try:
        box = bbox.extents([insert])
        if box.has_data:
            return box
    except Exception:
        pass
    return None


def _looks_like_drawing_number(text: str) -> bool:
    """True if *text* (after squashing whitespace) looks like a drawing number."""
    squashed = re.sub(r"\s+", "", text)
    return bool(
        _DRAWING_NUMBER_RE.match(squashed)
        and any(ch.isdigit() for ch in squashed)
    )


def _pair_number_and_name(
    hits: List[Tuple[float, float, str]],
) -> Tuple[str, str]:
    """
    From texts inside the search rectangle, identify Drawing Number and Name.

    The Drawing Number is normalised via _normalize_drawing_number() so that
    "AA - 401" in a DWG and "AA-401" in the PDF merge onto the same row.
    """
    if not hits:
        return "", ""

    numbers: List[Tuple[float, float, str]] = []
    others:  List[Tuple[float, float, str]] = []
    for x, y, text in hits:
        if _looks_like_drawing_number(text):
            numbers.append((x, y, text))
        else:
            others.append((x, y, text))

    if numbers and others:
        numbers.sort(key=lambda h: -h[1])
        others.sort(key=lambda h: -h[1])
        return _normalize_drawing_number(numbers[0][2]), others[0][2]

    hits_sorted = sorted(hits, key=lambda h: -h[1])
    if len(hits_sorted) >= 2:
        return _normalize_drawing_number(hits_sorted[0][2]), hits_sorted[1][2]
    return _normalize_drawing_number(hits_sorted[0][2]), ""


def extract_dwg_data(target_dir: str, block_name: str) -> pd.DataFrame:
    """
    Walk *target_dir* and extract Drawing Number / Drawing Name from every
    INSERT matching *block_name* in the host Model Space.
    """
    target_block = block_name.strip().lower()
    rows: List[dict] = []

    cad_files: List[str] = []
    for pat in ("*.dwg", "*.DWG", "*.dxf", "*.DXF"):
        cad_files.extend(glob.glob(os.path.join(target_dir, pat)))
    cad_files = sorted(set(cad_files))

    empty = pd.DataFrame(columns=["Source File", "Drawing Number", "Drawing Name"])
    if not cad_files:
        print(f"[WARN] No .dwg / .dxf files found in {target_dir}")
        return empty

    prev_cwd = os.getcwd()
    try:
        os.chdir(target_dir)
        print(f"[CAD ] Working directory : {os.getcwd()}")
        print(f"[CAD ] Files discovered  : {len(cad_files)}")

        for idx, full_path in enumerate(cad_files, 1):
            fname = os.path.basename(full_path)
            print(f"[CAD ] ({idx}/{len(cad_files)}) Processing {fname} ...",
                  end=" ", flush=True)

            try:
                doc = _load_doc(Path(fname))
            except Exception as exc:
                print(f"FAILED to open ({exc})")
                continue

            try:
                msp = doc.modelspace()
                tb_inserts = [
                    ins for ins in msp.query("INSERT")
                    if ins.dxf.name.strip().lower() == target_block
                ]
                if not tb_inserts:
                    print(f"no '{block_name}' in model space")
                    continue

                text_entities = list(msp.query("TEXT MTEXT"))
                file_rows = 0

                for tb in tb_inserts:
                    box = _compute_insert_bbox(tb)
                    if box is None:
                        continue  # unresolved xref — skip

                    min_x = float(box.extmin.x)
                    min_y = float(box.extmin.y)
                    max_x = float(box.extmax.x)
                    max_y = float(box.extmax.y)
                    width  = max_x - min_x
                    height = max_y - min_y
                    if width <= 0 or height <= 0:
                        continue

                    sx_max = max_x
                    sx_min = max_x - width  * X_RATIO
                    sy_min = min_y
                    sy_max = min_y + height * Y_RATIO

                    hits: List[Tuple[float, float, str]] = []
                    for ent in text_entities:
                        pt = _entity_point(ent)
                        if pt is None:
                            continue
                        x, y = pt
                        if sx_min <= x <= sx_max and sy_min <= y <= sy_max:
                            content = _entity_text(ent)
                            if content:
                                hits.append((x, y, content))

                    if not hits:
                        continue

                    dwg_no, dwg_name = _pair_number_and_name(hits)
                    if dwg_no or dwg_name:
                        rows.append({
                            "Source File":    fname,
                            "Drawing Number": dwg_no,
                            "Drawing Name":   dwg_name,
                        })
                        file_rows += 1

                print(f"Done ({file_rows} drawings, "
                      f"{len(tb_inserts)} title-block instance(s))")

            except Exception as exc:
                print(f"FAILED (unexpected error: {exc})")
                traceback.print_exc()

    finally:
        os.chdir(prev_cwd)

    df = pd.DataFrame(rows, columns=["Source File", "Drawing Number", "Drawing Name"])
    if not df.empty:
        df = (
            df[df["Drawing Number"] != ""]
            .drop_duplicates(subset=["Drawing Number"])
            .reset_index(drop=True)
        )
    print(f"[CAD ] Total unique drawings extracted: {len(df)}")
    return df


# ============================================================================
# 4.  COMPARE & WRITE EXCEL REPORT
# ============================================================================
def build_report(pdf_df: pd.DataFrame, dwg_df: pd.DataFrame, out_path: str) -> None:
    """Outer-merge on Drawing Number and write a highlighted report.xlsx."""
    pdf = pdf_df.rename(columns={
        "Drawing Name": "Drawing Name (PDF)",
        "Scale":        "Scale (PDF)",
    })
    dwg = dwg_df.rename(columns={
        "Drawing Name": "Drawing Name (DWG)",
    })

    merged = pdf.merge(dwg, on="Drawing Number", how="outer", indicator=True)

    for col in ["Drawing Name (PDF)", "Drawing Name (DWG)", "Scale (PDF)", "Source File"]:
        if col not in merged.columns:
            merged[col] = ""

    merged["Match Status"] = merged["_merge"].map({
        "both":       "MATCHED",
        "left_only":  "DWG MISSING",
        "right_only": "PDF MISSING",
    })
    merged = merged[[
        "Drawing Number",
        "Drawing Name (PDF)",
        "Drawing Name (DWG)",
        "Scale (PDF)",
        "Source File",
        "Match Status",
    ]].fillna("")

    merged.to_excel(out_path, index=False)

    red = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    wb  = load_workbook(out_path)
    ws  = wb.active

    h    = {cell.value: cell.column for cell in ws[1]}
    c_no = h["Drawing Number"]
    c_np = h["Drawing Name (PDF)"]
    c_nd = h["Drawing Name (DWG)"]
    c_sc = h["Scale (PDF)"]
    c_st = h["Match Status"]

    for row in range(2, ws.max_row + 1):
        status   = ws.cell(row=row, column=c_st).value or ""
        name_pdf = (ws.cell(row=row, column=c_np).value or "").strip()
        name_dwg = (ws.cell(row=row, column=c_nd).value or "").strip()
        scale    = (ws.cell(row=row, column=c_sc).value or "").strip()

        if status in ("DWG MISSING", "PDF MISSING"):
            for c in (c_no, c_np, c_nd, c_sc, c_st):
                ws.cell(row=row, column=c).fill = red
            continue

        if name_pdf.lower() != name_dwg.lower():
            ws.cell(row=row, column=c_np).fill = red
            ws.cell(row=row, column=c_nd).fill = red

        if scale == "":
            ws.cell(row=row, column=c_sc).fill = red

    wb.save(out_path)
    print(f"[XLSX] Report saved: {out_path}")


# ============================================================================
# 5.  MAIN
# ============================================================================
def main() -> None:
    target_dir, pdf_path, block_name = prompt_inputs()

    print("-" * 72)
    print(f"[INFO] Target dir : {target_dir}")
    print(f"[INFO] PDF path   : {pdf_path}")
    print(f"[INFO] Block name : {block_name}")
    print("-" * 72)

    # Resolve & configure the ODA File Converter eagerly, but only when the
    # target directory actually contains .dwg files — avoids prompting the
    # user for a converter path they do not need.
    dwg_hits = (
        glob.glob(os.path.join(target_dir, "*.dwg"))
        + glob.glob(os.path.join(target_dir, "*.DWG"))
    )
    if dwg_hits:
        _configure_odafc()

    out_path = os.path.abspath(REPORT_NAME)

    try:
        pdf_df = extract_pdf_table(pdf_path)
    except Exception as exc:
        print(f"[FATAL] PDF extraction aborted: {exc}")
        sys.exit(1)

    try:
        dwg_df = extract_dwg_data(target_dir, block_name)
    except Exception as exc:
        print(f"[FATAL] DWG extraction aborted: {exc}")
        traceback.print_exc()
        sys.exit(1)

    if pdf_df.empty and dwg_df.empty:
        print("[ERROR] Both datasets are empty. Nothing to compare.")
        sys.exit(1)

    try:
        build_report(pdf_df, dwg_df, out_path)
    except Exception as exc:
        print(f"[FATAL] Report generation failed: {exc}")
        traceback.print_exc()
        sys.exit(1)

    print("-" * 72)
    print("[DONE] All tasks completed successfully.")


if __name__ == "__main__":
    main()
