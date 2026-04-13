"""
drawing_list_validator.py  —  도면목록표(DWG) 기준 도면 검증 도구  (V1)
============================================================================
도면목록표.dwg 파일을 기준(Source of Truth)으로 삼아, 해당 목록에 등록된
모든 도면의 "도면명 / 도면번호 / 축척"이 실제 개별 DWG 파일의 도곽 내용과
일치하는지 자동으로 교차 검증하고 엑셀 리포트로 출력합니다.

기존 app.py (PDF ↔ CAD V46)가 PDF를 기준으로 사용하는 것과 달리,
본 도구는 CAD 자체에서 작성된 도면목록표 DWG를 기준으로 사용합니다.

실행 흐름
----------
1) 도면목록표 DWG 파싱
   - 모델스페이스/페이퍼스페이스의 TEXT/MTEXT 엔티티를 행·열로 군집화
   - "도면번호/도면명/축척" 헤더 행을 자동 인식해 컬럼 X 범위 계산
   - 헤더 아래 데이터 행을 컬럼별로 분배하여 목록 테이블 생성
2) 검증 대상 DWG 폴더 파싱 (app.extract_dwg_data 재사용)
3) 도면번호를 KEY로 병합하여 도면명/축척/누락 여부를 교차 검증
4) 불일치 셀·누락 행을 하이라이팅한 엑셀 리포트 저장
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 기존 V46 유틸을 재사용
from app import (
    _cad_문서_로드,
    _도면번호_세척,
    _extract_drawing_number,
    _축척_텍스트_정리,
    _텍스트_데이터_추출,
    extract_dwg_data,
)

# ============================================================================
# 전역 설정
# ============================================================================
리포트_이름: str = "도면목록표_검증리포트.xlsx"
DEBUG: bool = True

# 헤더 인식 키워드 (공백/대소문자 제거 후 부분 일치)
HDR_NO_KEYS: List[str] = [
    "도면번호", "도면 번호", "도서번호", "도번",
    "DWGNO", "DWG.NO", "DWG NO.", "DRAWINGNO",
]
HDR_NAME_KEYS: List[str] = [
    "도면명", "도면 명", "도면명칭", "도면 명칭",
    "도면이름", "도명", "제목", "TITLE", "DRAWINGNAME",
]
HDR_SCALE_KEYS: List[str] = ["축척", "SCALE", "스케일"]


def _normalize(s: str) -> str:
    """공백 제거 + 대문자 변환으로 표준화."""
    return re.sub(r"\s+", "", s or "").upper()


def _contains_any(text: str, keys: List[str]) -> bool:
    t = _normalize(text)
    return any(_normalize(k) in t for k in keys)


# ============================================================================
# 1. 도면목록표 DWG 파싱
# ============================================================================
def _iter_layouts(doc):
    """모델스페이스 + 페이퍼스페이스 레이아웃을 모두 순회."""
    try:
        names = list(doc.layout_names())
    except Exception:
        return [doc.modelspace()]
    layouts = []
    for name in names:
        try:
            layouts.append(doc.layout(name))
        except Exception:
            continue
    return layouts or [doc.modelspace()]


def _collect_layout_texts(layout) -> List[Tuple[float, float, str]]:
    """레이아웃에서 TEXT/MTEXT의 (x, y, 문자열) 목록 수집."""
    texts: List[Tuple[float, float, str]] = []
    for ent in layout.query("TEXT MTEXT"):
        d = _텍스트_데이터_추출(ent)
        if d and d[2]:
            texts.append(d)
    return texts


def _estimate_row_tolerance(texts: List[Tuple[float, float, str]]) -> float:
    """Y축 간격 분포를 분석해 행 병합 허용 오차를 자동 결정."""
    if len(texts) < 5:
        return 1.0
    ys = sorted({round(t[1], 3) for t in texts})
    gaps = [ys[i + 1] - ys[i] for i in range(len(ys) - 1) if ys[i + 1] - ys[i] > 0]
    if not gaps:
        return 1.0
    gaps.sort()
    # 하위 25% 간격 평균의 1.5배를 행 허용 오차로 사용
    small = gaps[: max(1, len(gaps) // 4)]
    return max(0.5, (sum(small) / len(small)) * 1.5)


def _group_into_rows(
    texts: List[Tuple[float, float, str]], y_tol: float
) -> List[List[Tuple[float, float, str]]]:
    """Y좌표 기반으로 텍스트를 행 단위로 군집화하고 각 행은 X로 정렬."""
    sorted_texts = sorted(texts, key=lambda t: -t[1])
    rows: List[List[Tuple[float, float, str]]] = []
    current: List[Tuple[float, float, str]] = []
    current_y: Optional[float] = None
    for t in sorted_texts:
        if current_y is None:
            current = [t]
            current_y = t[1]
        elif abs(current_y - t[1]) <= y_tol:
            current.append(t)
        else:
            rows.append(sorted(current, key=lambda x: x[0]))
            current = [t]
            current_y = t[1]
    if current:
        rows.append(sorted(current, key=lambda x: x[0]))
    return rows


def _find_header_rows(
    rows: List[List[Tuple[float, float, str]]],
) -> List[Tuple[int, Dict[str, Tuple[float, float]]]]:
    """헤더 행의 인덱스와 각 컬럼의 X 범위(low, high)를 반환."""
    results: List[Tuple[int, Dict[str, Tuple[float, float]]]] = []
    for idx, row in enumerate(rows):
        anchors: Dict[str, float] = {}
        for tx, _ty, txt in row:
            if "no" not in anchors and _contains_any(txt, HDR_NO_KEYS):
                anchors["no"] = tx
            elif "name" not in anchors and _contains_any(txt, HDR_NAME_KEYS):
                anchors["name"] = tx
            elif "scale" not in anchors and _contains_any(txt, HDR_SCALE_KEYS):
                anchors["scale"] = tx
        if len(anchors) < 3:
            continue
        # X 기준으로 컬럼 순서 정렬 후 인접 컬럼 중앙값을 경계로 사용
        ordered = sorted(anchors.items(), key=lambda kv: kv[1])
        ranges: Dict[str, Tuple[float, float]] = {}
        for i, (key, x) in enumerate(ordered):
            prev_x = ordered[i - 1][1] if i > 0 else x - 1e9
            next_x = ordered[i + 1][1] if i + 1 < len(ordered) else x + 1e9
            low = (prev_x + x) / 2 if i > 0 else -1e12
            high = (x + next_x) / 2 if i + 1 < len(ordered) else 1e12
            ranges[key] = (low, high)
        results.append((idx, ranges))
    return results


def _assign_column(tx: float, ranges: Dict[str, Tuple[float, float]]) -> Optional[str]:
    for col, (lo, hi) in ranges.items():
        if lo <= tx < hi:
            return col
    return None


def _clean_list_scale(raw: str) -> str:
    """도면목록표의 축척 컬럼 텍스트를 정규화.

    _축척_텍스트_정리는 '1/100' 형태 전용이므로, 목록표에서 종종 나타나는
    '100' 같은 순수 숫자나 'A=1/100' 같은 접두사를 처리할 수 있도록 보완.
    """
    if not raw:
        return "X"
    t = raw.strip()
    upper = t.upper()
    if "NONE" in upper or "N/A" in upper:
        return "NONE"
    # 접두사/접미사 제거 시도 후 기존 로직 적용
    cleaned = _축척_텍스트_정리(t)
    if cleaned != "X":
        return cleaned
    m = re.search(r"(\d{1,4})", t)
    if m:
        return f"1/{m.group(1)}"
    return "X"


def extract_drawing_list_from_dwg(dwg_path: str) -> pd.DataFrame:
    """도면목록표 DWG에서 (도면번호, 도면명, 축척) 목록을 추출."""
    print(f"[LIST] 도면목록표 분석 시작: {os.path.basename(dwg_path)}")
    doc = _cad_문서_로드(Path(dwg_path))

    데이터: List[Dict[str, str]] = []
    for layout in _iter_layouts(doc):
        texts = _collect_layout_texts(layout)
        if not texts:
            continue

        y_tol = _estimate_row_tolerance(texts)
        rows = _group_into_rows(texts, y_tol)
        headers = _find_header_rows(rows)
        if not headers:
            continue

        print(
            f"[LIST] 레이아웃 '{layout.name}' — 행 {len(rows)}개, "
            f"헤더 {len(headers)}개 감지 (y_tol={y_tol:.2f})"
        )

        for h_i, (hdr_idx, ranges) in enumerate(headers):
            end_idx = headers[h_i + 1][0] if h_i + 1 < len(headers) else len(rows)
            for row in rows[hdr_idx + 1 : end_idx]:
                buckets: Dict[str, List[Tuple[float, str]]] = {
                    "no": [], "name": [], "scale": [],
                }
                for tx, _ty, txt in row:
                    col = _assign_column(tx, ranges)
                    if col is None:
                        continue
                    buckets[col].append((tx, txt))

                def _join(key: str) -> str:
                    return " ".join(
                        t for _, t in sorted(buckets[key], key=lambda x: x[0])
                    ).strip()

                번호_raw = _join("no")
                명칭 = _join("name")
                축척_raw = _join("scale")

                if not (번호_raw and 명칭):
                    continue
                # 헤더가 연속 반복되는 경우(페이지 분할 등) 건너뛰기
                if (
                    _contains_any(번호_raw, HDR_NO_KEYS)
                    or _contains_any(명칭, HDR_NAME_KEYS)
                ):
                    continue

                extracted = _extract_drawing_number(번호_raw)
                if not extracted:
                    continue
                번호 = _도면번호_세척(extracted)
                축척 = _clean_list_scale(축척_raw)

                데이터.append(
                    {
                        "도면번호(LIST)": 번호,
                        "도면명(LIST)": 명칭,
                        "축척(LIST)": 축척,
                    }
                )
                if DEBUG:
                    print(f"    [DBG] LIST 행: {번호} | {명칭} | {축척}")

    df = pd.DataFrame(데이터)
    if not df.empty:
        df = df.drop_duplicates(subset=["도면번호(LIST)"]).reset_index(drop=True)
    print(f"[LIST] 총 {len(df)}개 도면 항목 추출")
    return df


# ============================================================================
# 2. 교차 검증 리포트 생성
# ============================================================================
def _pick_dwg_scale(row) -> str:
    """DWG 도곽에서 추출한 A1/A3 축척 중 유효한 값을 우선 선택."""
    a1 = str(row.get("축척_A1(DWG)", "") or "").strip()
    a3 = str(row.get("축척_A3(DWG)", "") or "").strip()
    if a1 and a1 not in ("X", "NONE"):
        return a1
    if a3 and a3 not in ("X", "NONE"):
        return a3
    return a1 or a3 or "X"


def build_validation_report(
    list_df: pd.DataFrame, dwg_df: pd.DataFrame, out_path: str
) -> None:
    list_df = list_df.copy()
    dwg_df = dwg_df.copy()

    if not dwg_df.empty:
        dwg_df["축척(DWG)"] = dwg_df.apply(_pick_dwg_scale, axis=1)
    else:
        dwg_df = pd.DataFrame(
            columns=[
                "파일명", "도면번호(DWG)", "도면명(DWG)",
                "축척_A1(DWG)", "축척_A3(DWG)", "축척(DWG)",
            ]
        )

    list_df["KEY"] = list_df.get("도면번호(LIST)", pd.Series(dtype=str)).astype(str).str.replace(" ", "")
    dwg_df["KEY"] = dwg_df.get("도면번호(DWG)", pd.Series(dtype=str)).astype(str).str.replace(" ", "")

    merged = pd.merge(list_df, dwg_df, on="KEY", how="outer", indicator=True)

    def _status(r) -> str:
        m = r["_merge"]
        if m == "left_only":
            return "DWG 누락"
        if m == "right_only":
            return "LIST 누락"
        mismatches: List[str] = []
        if _normalize(str(r.get("도면번호(LIST)", ""))) != _normalize(str(r.get("도면번호(DWG)", ""))):
            mismatches.append("도면번호")
        if str(r.get("도면명(LIST)", "")).strip() != str(r.get("도면명(DWG)", "")).strip():
            mismatches.append("도면명")
        if _normalize(str(r.get("축척(LIST)", ""))) != _normalize(str(r.get("축척(DWG)", ""))):
            mismatches.append("축척")
        return "일치" if not mismatches else "불일치: " + ",".join(mismatches)

    merged["검증결과"] = merged.apply(_status, axis=1)

    컬럼순서 = [
        "도면번호(LIST)", "도면명(LIST)", "축척(LIST)",
        "도면번호(DWG)", "도면명(DWG)", "축척(DWG)",
        "파일명", "검증결과",
    ]
    for c in 컬럼순서:
        if c not in merged.columns:
            merged[c] = ""
    결과 = merged[컬럼순서].fillna("")

    # 기준(LIST) 도면번호 → 일치 → 누락 순으로 보기 좋게 정렬
    정렬우선 = {"일치": 0}
    결과["_정렬"] = 결과["검증결과"].apply(
        lambda s: 정렬우선.get(s, 2 if s.startswith("불일치") else 1)
    )
    결과 = 결과.sort_values(
        by=["_정렬", "도면번호(LIST)", "도면번호(DWG)"]
    ).drop(columns=["_정렬"]).reset_index(drop=True)

    결과.to_excel(out_path, index=False)

    # ------- 하이라이팅 -------
    빨강 = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    노랑 = PatternFill(start_color="FFFFF2A6", end_color="FFFFF2A6", fill_type="solid")
    헤더색 = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

    wb = load_workbook(out_path)
    ws = wb.active
    ws.title = "검증결과"

    header_map = {cell.value: cell.column for cell in ws[1]}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = 헤더색

    widths = [18, 40, 12, 18, 40, 12, 28, 24]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    for row in range(2, ws.max_row + 1):
        상태 = ws.cell(row=row, column=header_map["검증결과"]).value or ""
        if 상태 in ("DWG 누락", "LIST 누락"):
            for c in header_map.values():
                ws.cell(row=row, column=c).fill = 노랑
            continue
        if 상태.startswith("불일치"):
            if "도면번호" in 상태:
                ws.cell(row=row, column=header_map["도면번호(LIST)"]).fill = 빨강
                ws.cell(row=row, column=header_map["도면번호(DWG)"]).fill = 빨강
            if "도면명" in 상태:
                ws.cell(row=row, column=header_map["도면명(LIST)"]).fill = 빨강
                ws.cell(row=row, column=header_map["도면명(DWG)"]).fill = 빨강
            if "축척" in 상태:
                ws.cell(row=row, column=header_map["축척(LIST)"]).fill = 빨강
                ws.cell(row=row, column=header_map["축척(DWG)"]).fill = 빨강

    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"[XLSX] 리포트 저장 완료: {out_path}")


# ============================================================================
# 3. 메인
# ============================================================================
def main() -> None:
    print("=" * 72)
    print(" 도면목록표(DWG) 기준 도면 검증 도구 V1")
    print("=" * 72)

    목록표경로 = input("1. 도면목록표 DWG 파일 경로: ").strip().strip('"')
    캐드경로 = input("2. 검증 대상 DWG 폴더 경로: ").strip().strip('"')
    블록이름 = input("3. 도곽 블록 이름: ").strip()
    try:
        base_w = float(input("4. 도곽 원본의 가로 길이 (예: 841): ").strip())
        base_h = float(input("5. 도곽 원본의 세로 길이 (예: 594): ").strip())
    except ValueError:
        base_w, base_h = 841.0, 594.0

    list_df = extract_drawing_list_from_dwg(목록표경로)
    if list_df.empty:
        print(
            "[WARN] 도면목록표에서 항목을 찾지 못했습니다. "
            "헤더(도면번호/도면명/축척) 표기를 확인하세요."
        )

    dwg_df = extract_dwg_data(캐드경로, 블록이름, base_w, base_h)

    out_path = os.path.abspath(리포트_이름)
    build_validation_report(list_df, dwg_df, out_path)

    print("-" * 72)
    print("[DONE] 작업이 완료되었습니다.")


if __name__ == "__main__":
    main()
